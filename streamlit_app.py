import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import io


def process_excel_file(uploaded_file):
    workbook = load_workbook(filename=uploaded_file)

    if '總表' not in workbook.sheetnames:
        raise ValueError("文件中不存在名為 '總表' 的工作表")

    summary_sheet = workbook['總表']
    dates = []
    for cell in summary_sheet['E2:ZZ2'][0]:
        if cell.value is not None:
            dates.append(cell.value)
        else:
            break

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for date_col_idx, date in enumerate(dates, start=5):
        date_str = date.strftime('%Y-%m-%d')
        if date_str not in workbook.sheetnames:
            workbook.create_sheet(title=date_str)

        date_sheet = workbook[date_str]
        date_sheet['A1'] = date_str
        date_sheet['A2'] = '總進度'
        progress_cell = summary_sheet.cell(row=3, column=date_col_idx)
        date_sheet['B2'] = progress_cell.value

        date_sheet['C1'] = summary_sheet['C1'].value
        date_sheet['D1'] = summary_sheet['D1'].value
        date_sheet['A3'].fill = light_blue_fill
        date_sheet['A3'] = '本日所有工作'

        for cell in ['A6', 'B6', 'C6']:
            date_sheet[cell] = '工作項目' if cell == 'A6' else '本日進度' if cell == 'B6' else '累積進度'
            date_sheet[cell].fill = yellow_fill
            date_sheet[cell].border = thin_border

        projects_with_progress = []
        for row_idx in range(5, summary_sheet.max_row + 1):
            project = summary_sheet.cell(row=row_idx, column=1).value
            if not project:
                continue

            daily_progress = summary_sheet.cell(row=row_idx, column=date_col_idx).value
            if daily_progress:
                projects_with_progress.append(project)
                new_row = len(date_sheet['A']) + 1
                date_sheet.cell(row=new_row, column=1).value = project
                date_sheet.cell(row=new_row, column=2).value = daily_progress

                total_progress = sum(
                    summary_sheet.cell(row=row_idx, column=col).value or 0
                    for col in range(5, date_col_idx + 1)
                )
                date_sheet.cell(row=new_row, column=3).value = total_progress

        date_sheet['A4'] = ', '.join(projects_with_progress)

    return workbook

st.title('Excel檔案處理器')

uploaded_file = st.file_uploader("請上傳xlsx檔案", type=['xlsx'])
if uploaded_file is not None:
    with st.spinner('正在處理...'):
        workbook = process_excel_file(uploaded_file)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        st.success('處理完成！')
        st.download_button(label="下載修改後的文件", data=output, file_name="修改後的文件.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", on_click=st.balloons)

